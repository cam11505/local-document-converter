"""Safe, deterministic XLSX adapter implemented with openpyxl."""

from __future__ import annotations

from collections.abc import Iterable, Sequence
from datetime import date, datetime, time
from itertools import zip_longest
from pathlib import Path
from typing import Any
from xml.etree.ElementTree import ParseError as XmlParseError
from xml.etree.ElementTree import iterparse
from zipfile import BadZipFile, ZipFile

from openpyxl import load_workbook  # type: ignore[import-untyped]
from openpyxl.utils.exceptions import InvalidFileException  # type: ignore[import-untyped]

from local_document_converter.config import ExcelSettings
from local_document_converter.domain.models import (
    DocumentIR,
    DocumentWarning,
    HeadingBlock,
    SourceInfo,
    TableBlock,
)
from local_document_converter.exceptions import InputValidationError, ParseError
from local_document_converter.parsers.base import ParseContext, ParserCapability

_SPREADSHEET_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"


class ExcelParser:
    """Convert an XLSX workbook into one heading and one table per worksheet."""

    capability = ParserCapability(
        name="openpyxl",
        supported_extensions=frozenset({".xlsx"}),
    )

    def __init__(self, settings: ExcelSettings | None = None) -> None:
        self._settings = settings or ExcelSettings()

    def parse(self, source: Path, context: ParseContext) -> DocumentIR:
        source = source.resolve()
        self._validate_source(source)
        data_only = self._bool_option(context, "data_only", self._settings.data_only)
        read_only = self._bool_option(context, "read_only", self._settings.read_only)
        max_rows = self._positive_int_option(
            context, "max_rows_per_sheet", self._settings.max_rows_per_sheet
        )
        max_columns = self._positive_int_option(
            context, "max_columns_per_sheet", self._settings.max_columns_per_sheet
        )

        values_workbook: Any = None
        formulas_workbook: Any = None
        try:
            values_workbook = load_workbook(
                source,
                read_only=read_only,
                data_only=data_only,
                keep_vba=False,
                keep_links=False,
            )
            if data_only:
                formulas_workbook = load_workbook(
                    source,
                    read_only=read_only,
                    data_only=False,
                    keep_vba=False,
                    keep_links=False,
                )

            blocks: list[HeadingBlock | TableBlock] = []
            warnings: list[DocumentWarning] = []
            formula_sheets = (
                list(formulas_workbook.worksheets) if formulas_workbook is not None else []
            )
            value_sheets = list(values_workbook.worksheets)
            if formulas_workbook is not None and (
                values_workbook.sheetnames != formulas_workbook.sheetnames
            ):
                raise ParseError("XLSX value and formula views have different worksheets")

            for sheet_index, value_sheet in enumerate(value_sheets):
                formula_sheet = formula_sheets[sheet_index] if formula_sheets else None
                sheet_name = str(value_sheet.title)
                worksheet_path = getattr(value_sheet, "_worksheet_path", None)
                merged_ranges = self._read_merged_ranges(source, worksheet_path)
                rows, start_row, start_column, sheet_warnings = self._materialize_sheet(
                    sheet_name=sheet_name,
                    value_rows=value_sheet.iter_rows(values_only=True),
                    formula_rows=(
                        formula_sheet.iter_rows(values_only=True)
                        if formula_sheet is not None
                        else None
                    ),
                    max_rows=max_rows,
                    max_columns=max_columns,
                )
                warnings.extend(sheet_warnings)
                if merged_ranges:
                    warnings.append(
                        DocumentWarning(
                            code="excel.merged_cells",
                            message=(
                                f"Worksheet '{sheet_name}' contains merged cells; "
                                "only each range's top-left value is retained"
                            ),
                            details={"sheet": sheet_name, "ranges": list(merged_ranges)},
                        )
                    )

                heading_order = len(blocks)
                blocks.append(
                    HeadingBlock(
                        id=f"sheet-{sheet_index + 1}-heading",
                        order=heading_order,
                        level=1,
                        text=sheet_name,
                        attributes={
                            "sheet_index": sheet_index,
                            "sheet_name": sheet_name,
                        },
                    )
                )
                width = len(rows[0]) if rows else 0
                blocks.append(
                    TableBlock(
                        id=f"sheet-{sheet_index + 1}-table",
                        order=heading_order + 1,
                        column_names=[_column_label(index) for index in range(1, width + 1)]
                        or None,
                        rows=rows,
                        attributes={
                            "sheet_index": sheet_index,
                            "sheet_name": sheet_name,
                            "start_column": start_column,
                            "start_row": start_row,
                        },
                    )
                )

            return DocumentIR(
                source=SourceInfo(
                    path=str(source),
                    media_type=(
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    ),
                    size_bytes=source.stat().st_size,
                ),
                blocks=blocks,
                warnings=warnings,
            )
        except InputValidationError:
            raise
        except ParseError:
            raise
        except (
            BadZipFile,
            InvalidFileException,
            KeyError,
            OSError,
            ValueError,
            XmlParseError,
        ) as exc:
            raise ParseError(f"could not parse XLSX workbook: {source}") from exc
        finally:
            if formulas_workbook is not None:
                formulas_workbook.close()
            if values_workbook is not None:
                values_workbook.close()

    @staticmethod
    def _validate_source(source: Path) -> None:
        if source.suffix.lower() != ".xlsx":
            raise InputValidationError("ExcelParser accepts only .xlsx files")
        if not source.exists():
            raise InputValidationError(f"input does not exist: {source}")
        if not source.is_file():
            raise InputValidationError(f"input is not a file: {source}")

    @staticmethod
    def _bool_option(context: ParseContext, name: str, default: bool) -> bool:
        value = context.options.get(name, default)
        if not isinstance(value, bool):
            raise InputValidationError(f"Excel option '{name}' must be a boolean")
        return value

    @staticmethod
    def _positive_int_option(context: ParseContext, name: str, default: int) -> int:
        value = context.options.get(name, default)
        if isinstance(value, bool) or not isinstance(value, int) or value <= 0:
            raise InputValidationError(f"Excel option '{name}' must be a positive integer")
        return value

    @classmethod
    def _materialize_sheet(
        cls,
        *,
        sheet_name: str,
        value_rows: Iterable[Sequence[object | None]],
        formula_rows: Iterable[Sequence[object | None]] | None,
        max_rows: int,
        max_columns: int,
    ) -> tuple[list[list[str | None]], int, int, list[DocumentWarning]]:
        raw_rows: list[list[str | None]] = []
        missing_formula_cells: list[str] = []
        formula_cells: set[tuple[int, int]] = set()
        formula_source: Iterable[Sequence[object | None]] = formula_rows or ()
        paired_rows = (
            zip_longest(value_rows, formula_source, fillvalue=())
            if formula_rows is not None
            else ((row, ()) for row in value_rows)
        )

        for row_number, (value_row, formula_row) in enumerate(paired_rows, start=1):
            if row_number > max_rows:
                raise InputValidationError(
                    f"worksheet '{sheet_name}' exceeds max_rows_per_sheet={max_rows}"
                )
            width = max(len(value_row), len(formula_row))
            if width > max_columns:
                raise InputValidationError(
                    f"worksheet '{sheet_name}' exceeds max_columns_per_sheet={max_columns}"
                )

            normalized: list[str | None] = []
            for column_index in range(width):
                value = value_row[column_index] if column_index < len(value_row) else None
                formula = formula_row[column_index] if column_index < len(formula_row) else None
                if isinstance(formula, str) and formula.startswith("="):
                    formula_cells.add((row_number - 1, column_index))
                    if value is None:
                        missing_formula_cells.append(
                            f"{_column_label(column_index + 1)}{row_number}"
                        )
                normalized.append(cls._serialize_cell(value))
            raw_rows.append(normalized)

        rows, start_row, start_column, trimmed = cls._trim_blank_edges(
            raw_rows, occupied_cells=formula_cells
        )
        warnings: list[DocumentWarning] = []
        if missing_formula_cells:
            warnings.append(
                DocumentWarning(
                    code="excel.formula_cache_missing",
                    message=(
                        f"Worksheet '{sheet_name}' has formulas without cached values; "
                        "affected cells are blank because data_only=True"
                    ),
                    details={"cells": missing_formula_cells, "sheet": sheet_name},
                )
            )
        if not rows:
            warnings.append(
                DocumentWarning(
                    code="excel.empty_sheet",
                    message=f"Worksheet '{sheet_name}' contains no cell values",
                    details={"sheet": sheet_name},
                )
            )
        elif trimmed:
            warnings.append(
                DocumentWarning(
                    code="excel.blank_edges_trimmed",
                    message=f"Blank outer rows or columns were trimmed from '{sheet_name}'",
                    details={
                        "sheet": sheet_name,
                        "start_column": start_column,
                        "start_row": start_row,
                    },
                )
            )
        return rows, start_row, start_column, warnings

    @staticmethod
    def _trim_blank_edges(
        rows: list[list[str | None]],
        *,
        occupied_cells: set[tuple[int, int]] | None = None,
    ) -> tuple[list[list[str | None]], int, int, bool]:
        occupied_cells = occupied_cells or set()
        nonempty_rows = [
            index
            for index, row in enumerate(rows)
            if any(value is not None for value in row)
            or any(row_index == index for row_index, _ in occupied_cells)
        ]
        if not nonempty_rows:
            return [], 1, 1, bool(rows)

        first_row = nonempty_rows[0]
        last_row = nonempty_rows[-1]
        column_indexes = [
            column_index
            for row in rows[first_row : last_row + 1]
            for column_index, value in enumerate(row)
            if value is not None
        ]
        column_indexes.extend(
            column_index
            for row_index, column_index in occupied_cells
            if first_row <= row_index <= last_row
        )
        first_column = min(column_indexes)
        last_column = max(column_indexes)
        width = last_column - first_column + 1
        result = [
            (row + [None] * (last_column + 1 - len(row)))[first_column : last_column + 1]
            for row in rows[first_row : last_row + 1]
        ]
        trimmed = (
            first_row > 0
            or last_row < len(rows) - 1
            or first_column > 0
            or any(len(row) > width for row in rows[first_row : last_row + 1])
        )
        return result, first_row + 1, first_column + 1, trimmed

    @staticmethod
    def _serialize_cell(value: object | None) -> str | None:
        if value is None:
            return None
        if isinstance(value, bool):
            return "true" if value else "false"
        if isinstance(value, datetime):
            return value.isoformat(timespec="seconds")
        if isinstance(value, (date, time)):
            return value.isoformat()
        return str(value)

    @staticmethod
    def _read_merged_ranges(source: Path, worksheet_path: object) -> tuple[str, ...]:
        if not isinstance(worksheet_path, str):
            return ()
        merged_ranges: list[str] = []
        with ZipFile(source) as archive, archive.open(worksheet_path) as worksheet_xml:
            for _, element in iterparse(worksheet_xml, events=("end",)):
                if element.tag == f"{{{_SPREADSHEET_NS}}}mergeCell":
                    reference = element.attrib.get("ref")
                    if reference:
                        merged_ranges.append(reference)
                element.clear()
        return tuple(merged_ranges)


def _column_label(index: int) -> str:
    label = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        label = chr(65 + remainder) + label
    return label
